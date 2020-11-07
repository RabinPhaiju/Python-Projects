import openpyxl
import os

wb = openpyxl.load_workbook('example.xlsx')

# print(wb.properties)

# print(wb.sheetnames)


sheet = wb['Sheet1']

cell = sheet['c2']
print(cell.value)

# ------- Edit the cell ------------
sheet['c2'].value = 'rabin'
print(cell.value)

# columns read------------------
for i in range(1,sheet.max_column):
    print(i, sheet.cell(row = 1, column = i).value)
    

# values
for i in range(1, sheet.max_row):
    print(i, sheet.cell(row=i, column=7).value,sheet.cell(row=i, column=3).value)

 # ------- update ----------
 # try to update in new file, if the program have some bug
wb.save('example.xlsx')


wb.close()
