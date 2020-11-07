import openpyxl
from openpyxl.styles import Font
from copy import copy

# ----- new excel -----------
wb = openpyxl.Workbook()
sheet = wb['Sheet'] # sheet title

font = Font(sz = 14, bold = True, italic=True)

sheet['A1'].value = 'Name'
sheet['A1'].font = copy(font)
sheet['B1'].value = 'Location'
sheet['B1'].font = copy(font)
sheet['C1'].value = 'Roll'

for i in range(2,10):
    sheet.cell(row=i, column=1).value = 'rabin'
    
wb.save('edit.xlsx')

# ------------ new sheet -----------
sheet1 = wb.create_sheet('Sheet1')
print(wb.sheetnames)
sheet1['A1'].value = 'New Sheet'

wb.save('edit.xlsx')


