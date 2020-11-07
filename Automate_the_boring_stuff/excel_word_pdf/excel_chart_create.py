import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)

wb = openpyxl.load_workbook('example.xlsx')
wb1 = Workbook()
sheet = wb['Sheet1']

ws = wb1.active

rows = []
for i in range(1, sheet.max_column):
    row_list = []
    row_list.append(sheet.cell(row=i, column=8).value)
    row_list.append(sheet.cell(row=i, column=9).value)
    row_list.append(sheet.cell(row=i, column=10).value)
    rows.append(row_list)


for row in rows:
    ws.append(row)

chart = AreaChart()
chart.title = "Area Chart"
chart.style = 13
chart.x_axis.title = rows[0][1]
chart.y_axis.title = rows[0][2]


cats = Reference(ws, min_col=1, min_row=1, max_row=7)
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "A15")

wb1.save("from_file.xlsx")